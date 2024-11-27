import os
import json
import shutil
import glob
import aiohttp
import threading
import time
from datetime import datetime, timedelta
import requests
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes, MessageHandler, filters
from openpyxl import load_workbook
from flask import Flask, request

# Получение токена из переменной окружения
TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', '7514978498:AAF3uWbaKRRaUTrY6g8McYMVsJes1kL6hT4')
if not TOKEN:
    raise ValueError("Необходимо установить TELEGRAM_BOT_TOKEN в переменные окружения.")

ADMIN_IDS = [476571220,39897938]

# Файлы и директории
current_directory = os.path.dirname(os.path.abspath(__file__))
template_file = os.path.join(current_directory, "template.xlsx")
output_file = os.path.join(current_directory, "output.xlsx")
log_file = os.path.join(current_directory, "logs.json")
archive_dir = os.path.join(current_directory, "archive")

reset_interval = timedelta(hours=36)
last_reset_time = datetime.now()

temp_data = {}

QUESTIONS = [
   ("Введите Описание.", "description"),
    ("Введите количество.", "quantity"),
    ("Введите количество Диск отрезной 125х2.5мм", "disks_125"),
    ("Введите количество Диск отрезной 180х2.5мм", "disks_180"),
    ("Введите количество Диск шлифовальный d-125", "grinding_125"),
    ("Введите количество Диск шлифовальный d-180.", "grinding_180"),
    ("Введите количество электродов 3 мм.", "electrodes_3mm"),
    ("Введите количество электродов ЛЭЗ УОНИ 13/55 Д-2,5 мм.", "electrodes_uoni"),
]

# Flask-приложение для веб-сервера
app = Flask(__name__)

def log_action(username: str, success: bool):
    """Логирует действие пользователя."""
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")  # Формат даты: 2024-11-24
    day_of_week = now.strftime("%A")     # День недели: Monday
    time_str = now.strftime("%H:%M:%S")  # Формат времени: 10:45:20

    # Определяем статус
    status = "Успешно" if success else "Ошибка"

    # Загружаем текущие данные лога
    log_data = {}
    if os.path.exists(log_file):
        with open(log_file, "r", encoding="utf-8") as f:
            log_data = json.load(f)

    # Обновляем данные для текущей даты
    if date_str not in log_data:
        log_data[date_str] = {day_of_week: []}
    elif day_of_week not in log_data[date_str]:
        log_data[date_str][day_of_week] = []

    log_data[date_str][day_of_week].append({
        "username": username,
        "time": time_str,
        "status": status
    })

    # Сохраняем изменения
    with open(log_file, "w", encoding="utf-8") as f:
        json.dump(log_data, f, ensure_ascii=False, indent=2)

# Функции для архивации, очистки старых данных и обработки webhook

def archive_old_file():
    """Архивирует старый файл и сохраняет его в архивную папку."""
    if not os.path.exists(output_file):
        return

    if not os.path.exists(archive_dir):
        os.makedirs(archive_dir)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    archive_name = os.path.join(archive_dir, f"output_{timestamp}.xlsx")
    shutil.copy2(output_file, archive_name)

def clean_old_archives(retain_days: int = 7):
    """Удаляет старые архивы, старше retain_days дней."""
    cutoff_time = time.time() - (retain_days * 86400)  # 86400 секунд в сутках

    for file_path in glob.glob(os.path.join(archive_dir, "output_*.xlsx")):
        if os.path.getmtime(file_path) < cutoff_time:
            os.remove(file_path)

def clean_old_logs(retain_days: int = 30):
    """Удаляет старые логи."""
    if not os.path.exists(log_file):
        return

    with open(log_file, "r", encoding="utf-8") as f:
        log_data = json.load(f)

    cutoff_date = (datetime.now() - timedelta(days=retain_days)).strftime("%Y-%m-%d")
    filtered_data = {
        date: logs for date, logs in log_data.items() if date >= cutoff_date
    }

    with open(log_file, "w", encoding="utf-8") as f:
        json.dump(filtered_data, f, ensure_ascii=False, indent=2)

async def keep_alive():
    """Периодически пингует Telegram API, чтобы избежать разрывов соединения."""
    api_url = f"https://api.telegram.org/bot{TOKEN}/getMe"
    while True:
        async with aiohttp.ClientSession() as session:
            try:
                async with session.get(api_url) as response:
                    print(f"Telegram API ping: {response.status} - {datetime.now()}")
            except Exception as e:
                print(f"Error during ping: {e}")
        await asyncio.sleep(300)  # Пинг каждые 5 минут

def set_webhook():
    """Устанавливает webhook для бота."""
    webhook_url = 'https://yourdomain.com/your_webhook_path'  # Укажите ваш домен и путь
    response = requests.get(f'https://api.telegram.org/bot{TOKEN}/setWebhook?url={webhook_url}')
    print(response.json())  # Печатает результат установки webhook

@app.route('/your_webhook_path', methods=['POST'])
def webhook():
    """Получает обновления от Telegram через webhook."""
    json_str = request.get_data().decode("UTF-8")
    update = Update.de_json(json_str, application.bot)
    application.update_queue.put(update)
    return "OK", 200

def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    keyboard = [[InlineKeyboardButton("Добавить данные", callback_data="add_data")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    update.message.reply_text(
        "Привет! Я бот для работы с таблицами.\n"
        "Нажмите 'Добавить данные', чтобы начать.",
        reply_markup=reply_markup,
    )

# Основные функции бота
async def add_data(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    username = update.effective_user.username or f"user_{user_id}"
    temp_data[user_id] = {'state': 0, 'data': {}}
    await update.callback_query.answer()
    await update.callback_query.message.reply_text(QUESTIONS[0][0])
    log_action(username, success=True)

# Пример простой обработки сообщений
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    username = update.effective_user.username or f"user_{user_id}"

    if user_id not in temp_data:
        return

    state = temp_data[user_id]['state']
    if state >= len(QUESTIONS):
        return

    question, key = QUESTIONS[state]
    user_input = update.message.text

    if key != "description" and not user_input.isdigit():
        await update.message.reply_text("Пожалуйста, введите корректное число.")
        log_action(username, success=False)  # Логируем ошибку
        return

    temp_data[user_id]['data'][key] = user_input
    temp_data[user_id]['state'] += 1

    if temp_data[user_id]['state'] < len(QUESTIONS):
        next_question = QUESTIONS[temp_data[user_id]['state']][0]
        await update.message.reply_text(next_question)
    else:
        await show_done_button(update)

async def done(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    username = update.effective_user.username or f"user_{user_id}"

    if user_id not in temp_data or not temp_data[user_id]['data']:
        await update.callback_query.message.reply_text("Вы ещё не начали ввод данных.")
        log_action(username, success=False)
        return

    try:
        # Проверка наличия файла и его создание, если не существует
        if not os.path.exists(output_file):
            workbook = load_workbook(template_file)
            sheet = workbook.active
            workbook.save(output_file)
        else:
            workbook = load_workbook(output_file)
            sheet = workbook.active
    except FileNotFoundError:
        await update.callback_query.message.reply_text("Ошибка работы с файлом.")
        log_action(username, success=False)
        return

    # Заполнение данных
    row = sheet.max_row + 1
    for i, (key, value) in enumerate(temp_data[user_id]['data'].items(), start=1):
        sheet.cell(row=row, column=i, value=value)

    workbook.save(output_file)

    temp_data[user_id] = {'state': 0, 'data': {}}  # Сброс данных после завершения

    await update.callback_query.message.reply_text("Данные успешно добавлены.")
    log_action(username, success=True)

async def show_done_button(update: Update) -> None:
    keyboard = [[InlineKeyboardButton("Завершить", callback_data="done")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Все данные введены. Завершить?", reply_markup=reply_markup)

def main():
    application = Application.builder().token(TOKEN).build()

    # Обработчики
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CallbackQueryHandler(add_data, pattern="^add_data$"))
    application.add_handler(CallbackQueryHandler(done, pattern="^done$"))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    # Устанавливаем webhook
    set_webhook()

    # Запуск webhook
    application.run_webhook(
        listen="0.0.0.0",  # Слушаем все IP
        port=8443,         # Порт
        url_path="webhook_path",  # Путь для webhook
        webhook_url='https://tbot-1-k0fj.onrender.com/webhook_path',  # Полный URL webhook
    )

if __name__ == "__main__":
    main()
