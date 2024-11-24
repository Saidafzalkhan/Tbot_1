import os
import json
import threading
import time
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes,
    MessageHandler,
    filters,
)
from openpyxl import load_workbook

# Получение токена из переменной окружения
TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', '7514978498:AAF3uWbaKRRaUTrY6g8McYMVsJes1kL6hT4')
if not TOKEN:
    raise ValueError("Необходимо установить TELEGRAM_BOT_TOKEN в переменные окружения.")

ADMIN_IDS = [476571220]

current_directory = os.path.dirname(os.path.abspath(__file__))
template_file = os.path.join(current_directory, "template.xlsx")
output_file = os.path.join(current_directory, "output.xlsx")
log_file = os.path.join(current_directory, "logs.json")

reset_interval = timedelta(hours=36)
last_reset_time = datetime.now()

temp_data = {}

QUESTIONS = [
    ("Введите описание для товара/услуги.", "description"),
    ("Введите количество.", "quantity"),
    ("Введите количество дисков 125 мм.", "disks_125"),
    ("Введите количество дисков 180 мм.", "disks_180"),
    ("Введите количество шлифовальных кругов 125 мм.", "grinding_125"),
    ("Введите количество шлифовальных кругов 180 мм.", "grinding_180"),
    ("Введите количество электродов 3 мм.", "electrodes_3mm"),
    ("Введите количество электродов УОНИ.", "electrodes_uoni"),
]

# Функции логирования
def log_action(username: str, success: bool):
    """Логирует действие пользователя."""
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")  # Формат даты: 2024-11-24
    day_of_week = now.strftime("%A")     # День недели: Monday
    time_str = now.strftime("%H:%M:%S")  # Формат времени: 10:45:20

    # Определяем статус
    status = "Успешно" if success else "Ошибка"

    # Загружаем текущие данные лога
    log_data = {}
    if os.path.exists(log_file):
        with open(log_file, "r", encoding="utf-8") as f:
            log_data = json.load(f)

    # Обновляем данные для текущей даты
    if date_str not in log_data:
        log_data[date_str] = {day_of_week: []}
    elif day_of_week not in log_data[date_str]:
        log_data[date_str][day_of_week] = []

    log_data[date_str][day_of_week].append({
        "username": username,
        "time": time_str,
        "status": status
    })

    # Сохраняем изменения
    with open(log_file, "w", encoding="utf-8") as f:
        json.dump(log_data, f, ensure_ascii=False, indent=2)


def clean_old_logs(retain_days: int = 30):
    """Удаляет старые логи."""
    if not os.path.exists(log_file):
        return

    with open(log_file, "r", encoding="utf-8") as f:
        log_data = json.load(f)

    cutoff_date = (datetime.now() - timedelta(days=retain_days)).strftime("%Y-%m-%d")
    filtered_data = {
        date: logs for date, logs in log_data.items() if date >= cutoff_date
    }

    with open(log_file, "w", encoding="utf-8") as f:
        json.dump(filtered_data, f, ensure_ascii=False, indent=2)
async def send_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Отправляет файл пользователю."""
    user_id = update.effective_user.id

    # Проверяем, имеет ли пользователь доступ
    if user_id in ADMIN_IDS:
        try:
            # Проверяем существование файла
            if os.path.exists(output_file):
                await update.callback_query.answer()  # Закрываем анимацию ожидания
                with open(output_file, "rb") as file:
                    await context.bot.send_document(chat_id=update.effective_chat.id, document=file)
            else:
                await update.callback_query.message.reply_text("Файл не найден. Убедитесь, что он был создан.")
        except Exception as e:
            await update.callback_query.message.reply_text(f"Ошибка при отправке файла: {e}")
    else:
        await update.callback_query.message.reply_text("У вас нет прав для скачивания таблицы.")


async def send_log(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Отправляет лог файл админу."""
    user_id = update.effective_user.id
    if user_id in ADMIN_IDS:
        try:
            with open(log_file, "rb") as file:
                await context.bot.send_document(chat_id=update.effective_chat.id, document=file)
        except FileNotFoundError:
            await update.callback_query.message.reply_text("Лог файл не найден.")
        except Exception as e:
            await update.callback_query.message.reply_text(f"Ошибка при отправке лог файла: {e}")
    else:
        await update.callback_query.message.reply_text("У вас нет прав для скачивания лог файла.")
# Основные функции бота
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    keyboard = [[InlineKeyboardButton("Добавить данные", callback_data="add_data")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        "Привет! Я бот для работы с таблицами.\n"
        "Нажмите 'Добавить данные', чтобы начать.",
        reply_markup=reply_markup,
    )

async def show_done_button(update: Update) -> None:
    """Отображает кнопку 'Готово'."""
    keyboard = [[InlineKeyboardButton("Готово", callback_data="done")]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Проверяем источник сообщения
    if update.message:
        await update.message.reply_text("Нажмите 'Готово', чтобы завершить.", reply_markup=reply_markup)
    elif update.callback_query:
        await update.callback_query.message.reply_text("Нажмите 'Готово', чтобы завершить.", reply_markup=reply_markup)


async def add_data(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    username = update.effective_user.username or f"user_{user_id}"
    temp_data[user_id] = {'state': 0, 'data': {}}
    await update.callback_query.answer()
    await update.callback_query.message.reply_text(QUESTIONS[0][0])
    log_action(username, success=True)


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    username = update.effective_user.username or f"user_{user_id}"

    if user_id not in temp_data:
        return

    state = temp_data[user_id]['state']
    if state >= len(QUESTIONS):
        return

    question, key = QUESTIONS[state]
    user_input = update.message.text

    if key != "description" and not user_input.isdigit():
        await update.message.reply_text("Пожалуйста, введите корректное число.")
        log_action(username, success=False)  # Логируем ошибку
        return

    temp_data[user_id]['data'][key] = user_input
    temp_data[user_id]['state'] += 1

    if temp_data[user_id]['state'] < len(QUESTIONS):
        next_question = QUESTIONS[temp_data[user_id]['state']][0]
        await update.message.reply_text(next_question)
    else:
        await show_done_button(update)

async def restart_process(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Сбрасывает данные пользователя и начинает заново."""
    user_id = update.effective_user.id
    if user_id in temp_data:
        del temp_data[user_id]  # Очистить данные пользователя
    await add_data(update, context)  # Сбросить процесс и начать заново
async def done(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id
    username = update.effective_user.username or f"user_{user_id}"

    if user_id not in temp_data or not temp_data[user_id]['data']:
        await update.callback_query.message.reply_text("Вы ещё не начали ввод данных.")
        log_action(username, success=False)
        return

    try:
        # Проверка наличия файла и его создание, если не существует
        if not os.path.exists(output_file):
            workbook = load_workbook(template_file)
            sheet = workbook.active
            workbook.save(output_file)
        else:
            workbook = load_workbook(output_file)
            sheet = workbook.active
    except FileNotFoundError:
        await update.callback_query.message.reply_text("Шаблон таблицы не найден.")
        log_action(username, success=False)
        return

    data = temp_data[user_id]['data']
    row_index = 2

    # Поиск первой строки, где все ячейки пусты
    while any(sheet.cell(row=row_index, column=col).value is not None for col in range(2, sheet.max_column + 1)):
        row_index += 1

    row = [
        None,
        datetime.now().strftime("%d.%m.%Y г."),
        data.get('description', ""),
        None,
        data.get('quantity', ""),
        None,
        None,
        data.get('disks_125', ""),
        data.get('disks_180', ""),
        data.get('grinding_125', ""),
        data.get('grinding_180', ""),
        data.get('electrodes_3mm', ""),
        data.get('electrodes_uoni', ""),
    ]

    for col_index, value in enumerate(row, start=1):
        sheet.cell(row=row_index, column=col_index, value=value)

    try:
        workbook.save(output_file)
    except Exception as e:
        await update.callback_query.message.reply_text(f"Ошибка при сохранении файла: {e}")
        log_action(username, success=False)
        return
    finally:
        workbook.close()

    del temp_data[user_id]

    keyboard = [
        [InlineKeyboardButton("Скачать таблицу", callback_data="send_file")],
        [InlineKeyboardButton("Скачать лог", callback_data="send_log")],  # Кнопка для скачивания лог файла
        [InlineKeyboardButton("Начать заново", callback_data="restart")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.callback_query.message.reply_text("Таблица сформирована!", reply_markup=reply_markup)

    log_action(username, success=True)
def main():
    application = Application.builder().token(TOKEN).build()
    application.add_handler(CommandHandler("start", start))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    application.add_handler(CallbackQueryHandler(add_data, pattern="add_data"))
    application.add_handler(CallbackQueryHandler(done, pattern="done"))
    application.add_handler(CallbackQueryHandler(send_file, pattern="send_file"))
    application.add_handler(CallbackQueryHandler(send_log, pattern="send_log"))  # Обработчик для отправки лог файла
    application.add_handler(CallbackQueryHandler(restart_process, pattern="restart"))

    # Удаление старых логов при запуске
    clean_old_logs()

    application.run_polling()


if __name__ == "__main__":
    main()

